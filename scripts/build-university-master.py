#!/usr/bin/env python3
"""文部科学省『全国大学一覧』Excel から SHINROMii 用大学マスター JSON を生成する。

来年度は公式 Excel を差し替えて、このスクリプトを再実行すれば更新できる。
大学名・学部名の推測追加はしない。学部表に無い学部は収録しない。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl が必要です。python3 -m pip install openpyxl\n")
    raise

PREFECTURES = [
    "北海道",
    "青森県",
    "岩手県",
    "宮城県",
    "秋田県",
    "山形県",
    "福島県",
    "茨城県",
    "栃木県",
    "群馬県",
    "埼玉県",
    "千葉県",
    "東京都",
    "神奈川県",
    "新潟県",
    "富山県",
    "石川県",
    "福井県",
    "山梨県",
    "長野県",
    "岐阜県",
    "静岡県",
    "愛知県",
    "三重県",
    "滋賀県",
    "京都府",
    "大阪府",
    "兵庫県",
    "奈良県",
    "和歌山県",
    "鳥取県",
    "島根県",
    "岡山県",
    "広島県",
    "山口県",
    "徳島県",
    "香川県",
    "愛媛県",
    "高知県",
    "福岡県",
    "佐賀県",
    "長崎県",
    "熊本県",
    "大分県",
    "宮崎県",
    "鹿児島県",
    "沖縄県",
]

SECTION_STOP = {
    "研究科",
    "学部沿革",
    "大学院沿革",
    "専攻科",
    "別科",
    "通信教育",
    "沿革",
    "学習センター",
    "附属施設",
    "国立大学附置研究所",
    "大学沿革",
}

FACULTY_KEEP = re.compile(r"(学部|学群|学域|学環)")
TITLE_RE = re.compile(r"^(国立|公立|私立)\s+(.+)$")

SOURCE_FILES = [
    ("01.xlsx", "国立"),
    ("02.xlsx", "公立"),
    ("03-1.xlsx", "私立"),
    ("03-2.xlsx", "私立"),
    ("03-3.xlsx", "私立"),
    ("03-4.xlsx", "私立"),
    ("03-5.xlsx", "私立"),
    ("03-6.xlsx", "私立"),
    ("03-7.xlsx", "私立"),
    ("03-8.xlsx", "私立"),
    ("05.xlsx", "私立"),
]


def prefecture_from(text: object) -> str:
    if not text:
        return ""
    value = str(text)
    for prefecture in PREFECTURES:
        if prefecture in value:
            return prefecture
    return ""


def parse_title(text: object) -> tuple[str, str]:
    raw = str(text).replace("\u3000", " ").strip()
    match = TITLE_RE.match(raw)
    if not match:
        return "", raw
    name = re.sub(r"（[^）]*）\s*$", "", match.group(2).strip()).strip()
    return match.group(1), name


def is_faculty_name(name: str) -> bool:
    value = name.strip()
    if not value or value in {"学部", "学科", "都道府県", "名称"}:
        return False
    if "共同実施" in value or "沿革" in value:
        return False
    if "研究科" in value and "学部" not in value:
        return False
    if FACULTY_KEEP.search(value):
        return True
    return "学部等連係" in value


def faculty_id(school_code: str, faculty_name: str) -> str:
    slug = faculty_name.replace(" ", "").replace("\u3000", "")
    return f"shinromii-fac-{school_code}-{slug}"


def parse_sheet(ws, sheet_name: str) -> dict:
    rows = list(ws.iter_rows(min_row=1, max_row=500, max_col=14, values_only=True))
    university_type, name, code, address = "", sheet_name, "", ""
    faculties: list[str] = []

    for index, row in enumerate(rows):
        cell = row[1] if len(row) > 1 else None
        if index == 0 and cell:
            university_type, name = parse_title(cell)
        if cell == "学校コード" and index + 1 < len(rows):
            nxt = rows[index + 1]
            code = str(nxt[1] or "").strip()
            address = nxt[11] if len(nxt) > 11 else None
        if str(cell or "").strip() != "学部" or index + 1 >= len(rows):
            continue
        headers = [str(value).strip() if value else "" for value in rows[index + 1]]
        if "学科" not in headers:
            continue
        for later in rows[index + 2 :]:
            value = later[1] if len(later) > 1 else None
            if value in (None, ""):
                continue
            text = str(value).strip()
            if text in SECTION_STOP:
                break
            if is_faculty_name(text) and text not in faculties:
                faculties.append(text)
        break

    prefecture = prefecture_from(address)
    if not prefecture:
        for row in rows:
            for value in row:
                prefecture = prefecture_from(value)
                if prefecture:
                    break
            if prefecture:
                break

    return {
        "type": university_type,
        "name": name,
        "code": code,
        "prefecture": prefecture,
        "faculties": faculties,
    }


def build_master(source_dir: str) -> list[dict]:
    universities = []
    seen_codes: set[str] = set()

    for filename, default_type in SOURCE_FILES:
        path = os.path.join(source_dir, filename)
        if not os.path.exists(path):
            raise FileNotFoundError(f"公式 Excel が見つかりません: {path}")
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            parsed = parse_sheet(workbook[sheet_name], sheet_name)
            university_type = parsed["type"] or default_type
            code = parsed["code"]
            name = parsed["name"]
            if not code or not name:
                raise RuntimeError(f"{filename} / {sheet_name}: 学校コードまたは大学名を読めませんでした")
            if not re.fullmatch(r"F\d{12}", code):
                raise RuntimeError(f"{filename} / {sheet_name}: 学校コードの形が想定と違います: {code}")
            if code in seen_codes:
                raise RuntimeError(f"学校コードが重複しています: {code} ({name})")
            seen_codes.add(code)
            universities.append(
                {
                    "id": f"mext-{code}",
                    "schoolCode": code,
                    "name": name,
                    "type": university_type,
                    "prefecture": parsed["prefecture"],
                    "faculties": [
                        {"id": faculty_id(code, faculty_name), "name": faculty_name}
                        for faculty_name in parsed["faculties"]
                    ],
                }
            )
        workbook.close()

    universities.sort(key=lambda item: (item["prefecture"], item["name"]))
    return universities


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source-dir", required=True, help="文科省 Excel を置いたディレクトリ")
    parser.add_argument("--out", required=True, help="出力 JSON")
    parser.add_argument("--checked-at", default=date.today().isoformat())
    args = parser.parse_args()

    universities = build_master(args.source_dir)
    missing_pref = [item["name"] for item in universities if not item["prefecture"]]
    if missing_pref:
        raise RuntimeError("都道府県を読めない大学があります: " + ", ".join(missing_pref[:10]))

    payload = {
        "academicYear": "2025",
        "academicYearLabel": "令和7年度",
        "sourceName": "文部科学省『令和7年度全国大学一覧』",
        "sourceUrl": "https://www.mext.go.jp/a_menu/koutou/ichiran/mext_00050.html",
        "checkedAt": args.checked_at,
        "idNote": "大学idの mext-{学校コード} は文科省学校コードに基づく。学部idの shinromii-fac- はSHINROMii内部ID。",
        "universities": universities,
    }

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    with open(args.out, "w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")

    type_counts = {}
    for item in universities:
        type_counts[item["type"]] = type_counts.get(item["type"], 0) + 1
    faculty_count = sum(len(item["faculties"]) for item in universities)
    print(
        json.dumps(
            {
                "universities": len(universities),
                "faculties": faculty_count,
                "types": type_counts,
                "out": args.out,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
